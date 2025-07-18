#!/home/dave/Python/FIT/.venv/bin/python
# 04/??/2020: V1 - New
# 04/18/2020: V2 - Add code to create XLSX file; Remove code to create CSV file
# 10/14/2021: V2.1 - Modify code to see new Wahoo Speed sensor and changes from Elemnt upgrades
# 10/15/2021: V3.0 - Added code to send email from here (and removed from BASH script)
# 10/16/2021: v3.1 - Added stop date and time to text message with power levels and changed these msgs slightly
# 10/23/2021: V3.1 - Moved log write with program name and version to bottom of program and log file.
# 11/12/2021: V3.1 - Changed name of XLSX-file_name to have 00_ as first characters to sort at top of list
# 12/15/2021: V3.1 - Change recharge trigger level to 20%
# 02/27/2022: V3.2 - Added line to text message with ride miles and time
# 03/06/2022: V3.2 - Very minor change ('DB' to 'database') in a message
# 03/08/2022: V3.3 - Added code to check specific battery charge thresholds by device.
# 03/09/2022: V3.3 - Changed power-text message to be two parts - finished_msg & power_msg
# 03/11/2022: V3.35- Fixed a bug introduced in last change where battery level of the ELEMNT wasn't sent
# 03/18/2022: V3.4 - Added code to send a 'finished' email/text to Heather and me
# 03/20/2022: V3.5 - Added code to write XLSX file of today's ride to 'dow_Ride' in new folder (dow=3-char weekday) and
#                    added code to only send emails if ride date is today
# 03/22/2022: V3.5 - Minor message changes.
# 04/04/2022: V3.5 - Removed my id from receiving 'finished ride' text email; Moved these to top with other constants
# 04/06/2022: V3.5 - Added weekday to name of Excel file
# 06/04/2022: V3.6 - Added and modified code to use common config folder and modules
# 06/14/2022: V3.6 - Trimmed words from messages.
# 08/13/2022: V3.61- Added code to decrement shift counts by one.
# 08/19/2022: V3.62- Added code to import warnings and ignore user warnings I started to see after my system rebuild.
# 10/24/2022: V3.7 - Added and modified code to send battery level notify only if a sensor is installed.
# 12/31/2022: V3.71- Formatting changes for output messages sent via email and displayed
# 01/02/2023: V3.72- Added feet to ascent and descent report line
# 01/04/2023: V3.72- Changed KMH to KPH
# 01/26/2023: V3.73- Changed contents of some messages (combining on one line, removing a line)
# 01/31/2023: V3.74- Minor change to message with ride start, end, and duration
# 02/15/2023: V3.75- Changed to use /home/dave/tmp
# 11/03/2023: V3.80- Removed all references and messages about total ascent and descent.
#                    These values are no longer in the FIT file since the last update of the Elemnt from Wahoo
# 11/04/2023: V3.80- Corrected an error I let happen in yesterday's changes.  Also rename log file FITProc
# 11/07/2023: V3.90- Put Ascent and Descent variable back in but set to 0 if not found 
# 12/03/2023: V3.91- Changed MySQL connection & Email to use values directly from imported MySQL & Email credentials
# 2024-01-10: V3.92- Added line for unverified SSL context (replacing default) to bypass SSL errors when sending email
# 2024-02-02: V3.93- Changed suffix for Excel files to lower case letters
# 2024-03-04: V4.00- Changed to use openpyxl and update consolidated spreadsheet directly
# 2024-03-08: V4.00- Changed last row to scan when looking for an empty row in the spreadsheet
# 2024-03-08: V4.10- Changed key of 'ELEMNT' to 'ELEMNT ROAM'
# 2024-04-04: V4.20- Added lists for GPS and Eshift products; keys for battery charges are fixed at 'GPS' and 'Eshift'
# 2024-04-13: V4.20- Slight change to message addons to recharge or notify that speed sensor is dead.
# 2024-04-15: V4.21- Added code to set default font and font size in spreadsheet
# 2024-06-06: V4.3 - Added code to calculate and show in the log messages the YTD miles ridden
# 2024-07-02: V4.31- Minor change to a message text.
# 2024-07-03: V4.4 - Added dictionary for names of Eshift and GPS devices; Changed references to Eshift
# 2024-11-14: V4.41- Set default final charge value for Eshift device to -1 and show "N/A' in messages
# 2024-11-15: V4.5 - Create a string for Eshift battery % for messages; Show N/A if not found;
#                    Removed some code added yesterday
# 2024-12-25: V4.5 - Changed trigger battery percentage for Roam to 35%
# 2025-01-17: V4.6 - Changed trigger battery percentage for Roam to 25%; Added # of rides in a log message
#                  - Added code to not add ride to spreadsheet if it didn't occur this year
# 2025-01-27: V4.7 - Added code to exit if the FIT file as already in the database.  Removed check for this later on.
# 2025-02-04: V4.71- Changed battery trigger percentage for Roam to 30%
# 2025-05-01: V4.72- Changed format for reporting of number of records found to include a comma
# 2025-05-13: V4.80- Removed message about gear changes if none reported
# 2025-05-15: V4.90- 1) If gear changes are not found (set to -1), write 0 to database;
#                    2) Change email addresses due to AT&T change to eliminate ability to send texts via email
# 2025-05-16: V4.90- Fixed error caused by a bad variable name
# 2025-05-19: V4.91- Changed code to send emails only if email list or variable is not empty
# 2025-05-23: V4.92- Changed code to send battery status email if any device is marked as 'recharge'
# 2025-05-24: V4.92- Minor grammar change to message about the battery status email
# 2025-05-28: V5.00- Change email messages to be text messages via TextBelt SMS service
# 2025-05-29: V5.01- Added two newlines at the end of the text messages; Removed two print statements; Minor log message changes.
# 2025-05-30: V5.10- Added code to parse response from text server into log messages
# 2025-06-01: V5.11- Reversed decision to send text for battery status message - now it's going via email.  Uncommented email lines
# 2025-06-07: V5.12- Added contents of message sent to Heather via text to the log file
# 2025-06-08: V5.13- Slightly modified log message with contents of 'ride finished' text; Added date to log file name
# 2025-06-09: V5.14- More changes to log message with contents of 'ride finished' text & change to 'ride finished' message
# 2025-06-10: V5.15- Minor changes to 'ride finished' message
# 2025-06-11: V5.15- More minor changes to 'ride finished' message (Removed 'Rode', round to 10th of a mile, removed seconds on time)
# 2025-06-12: V5.15- More minor changes to 'ride finished' message (shortened dash separators, removed two new lines characters)
# 2025-06-22: V5.16- Changes to 'recharge message' for spacing.  Also, renamed some variables for consistancy
# 2025-07-07: V5.17- Change recharge trigger for Roam to 25%, DI2 to 20%; Change subject text for recharge email message.
# 2025-07-10: V5.18- Changed location of log file to 'Logs/FIT'
# 2025-07-11: N/C  - Changed DI2 recharge trigger limit to 25%
# 2025-07-18: V5.20- Changed to import values for text message receiver & API code

Program_Version = 5.20

import os
import sys
import warnings
warnings.simplefilter("ignore", UserWarning)
from datetime import date, datetime, timezone, timedelta
import pymysql      # For access to MySQL
import fitdecode    # For decoding FIT contents
import openpyxl     # To open and update Excel spreadsheet
from openpyxl.styles import DEFAULT_FONT
import smtplib      # For Email
import ssl          # Also for Email
import requests     # To send request for text message to service provider
sys.path.insert(0, '/home/dave/Python/_Configs')    # To get following imports from my library
import _MySQL       # MySQL login values
import _EMail       # Email login values
import _TextMsg     # Text message receiver and API code

# Common MySQL database values
Server = "localhost"
Database = "CyclingData"

# Email Receiver
recharge_msg_to = "dalexnagy@gmail.com"

# File names and folders
# Get current date and set other variables
today = date.today()
log_file_name = "/home/dave/Logs/FIT/{:4d}-{:02d}-{:02d}_FITProc_Log.txt".format(today.year, today.month, today.day)
XLSX_file_name = "/home/dave/Documents/Spreadsheets/BikeRides.xlsx"
worksheet_name = str(datetime.today().year)    # Default is this year

# Bike Sensor Flag
SpeedSensor_Installed_Flag = False
# Battery Recharge Flag
Battery_Recharge_Flag = False

# Lists of possible product names in FIT file
GPS_Products = ["ELEMNT", "ELEMNT ROAM"]
Eshift_Products = ["Shimano Di2"]
Device_Tags = {"GPS":"Roam", "Eshift":"DI2"}
# Recharge trigger thresholds
recharge_thresholds = [["GPS", 25], ["Eshift", 25]]
default_threshold = 20

# Dashes - Change value after * for length
dashes = '-' * 22
src_file = ""
src_folder = ""
# Testing
# Use any of next lines if testing with only one file - be sure to uncomment year if neede
# ----
#src_folder = "2024"
#src_file = "2024-07-01-103227-ELEMNT ROAM 2EFC-73-0.fit"      # A normal ride
#src_file = "2024-11-14-113554-ELEMNT ROAM 2EFC-163-0.fit"      # No DI2
#src_folder = "2025"
#src_file = "2025-05-28-104147-ELEMNT ROAM 5221-110-0.fit"    # Recent Ride & One device to be recharged
# And remember to reset the comment indicator!!

log_file = open(log_file_name, "w")

# log_file.write("{} Processing started\n".format(str(datetime.now().strftime("%H:%M:%S"))))

if len(src_file) == 0:
    # if src_file is not set to a value, search upload folder for files
    # Upload folder where FIT file is sent directly from Wahoo iPhone app
    src_dir = "/home/dave/Dropbox/apps/WahooFitness"
    files = os.listdir(src_dir)
    files.sort()
    fit_files = [file for file in files if file[-4:].lower() == '.fit']
    if len(fit_files) == 0:
        log_file.write("{} Processing started - NO files found in '"+src_dir+"' - Exiting\n"
                       .format(str(datetime.now().strftime("%H:%M:%S"))))
        print("NO files found in '"+src_dir+"' - Exiting")
        exit(1)
    else:
        log_file.write("{} Processing started for files in '{}'\n"
                       .format(str(datetime.now().strftime("%H:%M:%S")), src_dir))
else:
    src_dir = "/home/dave/WahooFitness/"
    if src_folder != "":
        src_dir = src_dir + src_folder
    fit_files = [src_file]
    log_file.write("{} Processing started for file: '{}/{}'\n"
                   .format(str(datetime.now().strftime("%H:%M:%S")), src_dir, src_file))

final_charge = {}
final_charge["Eshift"] = [datetime.now(), -1]
frame_names = ["device_info", "event", "session"]
speed_dev_names = ["Wahoo SPEED", "Speed"]

# Connect to database
db = pymysql.connect(host=Server, user=_MySQL.db_user, passwd=_MySQL.db_user_password, db=Database, autocommit=True)
cursor = db.cursor()
# Get number of rows in table
cursor.execute("SELECT COUNT(*) FROM RideStats")
# Fetch a single row using fetchone() method.
# log_file.write("{} Active database is '{}' on host '{}'\n".format(str(datetime.now().strftime("%H:%M:%S")),
# Database, Server))
log_file.write("{} Found {:,} ride records in '{}' on '{}'\n"
               .format(str(datetime.now().strftime("%H:%M:%S")), int(cursor.fetchone()[0]), Database, Server))

file_ctr = 0

for fit_file in fit_files:
    file_ctr += 1
    log_file.write("{} File #{:4d} - {}\n".format(str(datetime.now().strftime("%H:%M:%S")), file_ctr, fit_file))
    print("{} Processing File #{:4d} - {}".format(str(datetime.now().strftime("%H:%M:%S")), file_ctr, fit_file))
    sql = "SELECT FileName FROM RideStats WHERE FileName = '{}'".format(fit_file)
    cursor.execute(sql)

    # Get number of rows retrieved
    if cursor.rowcount > 0:
        log_file.write("{} File '{}' has been processed - No further action taken.\n"
                       .format(str(datetime.now().strftime("%H:%M:%S")), fit_file))
        print("{} File '{}' has been processed - No further action taken."
                       .format(str(datetime.now().strftime("%H:%M:%S")), fit_file))
        exit(0)

    full_fit_file = src_dir+"/"+fit_file

    # reset counters
    front_gear_changes = 0
    rear_gear_changes = 0
    frame_no = 0
    device_info_frames = 0
    Wahoo_Speed_frames = 0
    # Set values of variables that may not be in FIT file
    max_neg_grade_pct = 0
    max_pos_grade_pct = 0
    total_ascent_meters = 0
    total_descent_meters = 0

    # Process FIT file frames
    with fitdecode.FitReader(full_fit_file) as fit_frame:
        try:
            for frame in fit_frame:
                frame_no += 1

                if isinstance(frame, fitdecode.FitDataMessage):
                    try:
                        name_idx = frame_names.index(frame.name)
                    except ValueError:
                        name_idx = -1

                    if name_idx > -1:
                        # 'event' frame has gear change events reported by E-shift system
                        if frame.name == "event":
                            fields = frame.fields
                            for f in fields:
                                if f.value == "rear_gear_change":
                                    rear_gear_changes += 1
                                if f.value == "front_gear_change":
                                    front_gear_changes += 1

                        # 'session' frame has end-of-ride statistics
                        if frame.name == "session":
                            fields = frame.fields
                            for f in fields:
                                if f.name == "timestamp":
                                    stop_time = f.value
                                elif f.name == 'start_time':
                                    start_time = f.value
                                elif f.name == "total_elapsed_time":
                                    elapsed_time_secs = f.value
                                elif f.name == "total_timer_time":
                                    ride_time_secs = f.value
                                elif f.name == "avg_speed":
                                    avg_speed_meters = f.value
                                elif f.name == "max_speed":
                                    max_speed_meters = f.value
                                elif f.name == "total_distance":
                                    ride_distance_meters = f.value
                                elif f.name == "min_altitude":
                                    min_altitude_meters = f.value
                                elif f.name == "max_altitude":
                                    max_altitude_meters = f.value
                                elif f.name == "max_neg_grade":
                                    max_neg_grade_pct = f.value
                                elif f.name == "max_pos_grade":
                                    max_pos_grade_pct = f.value
                                elif f.name == "total_ascent":
                                    total_ascent_meters = f.value
                                elif f.name == "total_descent":
                                    total_descent_meters = f.value

                        # 'device_info' has battery charge information
                        if frame.name == "device_info":
                            fields = frame.fields
                            device_info_frames += 1
                            for flds in fields:
                                Dev_Charge_Str = ""
                                Description = ""
                                Product_Name = ""
                                for f in fields:
                                    if f.name == "charge":
                                        # print("charge f.value = ", f.value)
                                        if f.value is None:
                                            Dev_Charge_Str = "N/A"
                                        else:
                                            try:
                                                Dev_Charge_Str = str(f.value)
                                            except (TypeError, ValueError):
                                                Dev_Charge_Str = "????"
                                    elif f.name == 'descriptor':
                                        # print("descriptor f.value = ", f.value)
                                        if f.value is None:
                                            Description = "----"
                                        else:
                                            Description = f.value
                                    elif f.name == "product_name":
                                        # print("product_name f.value = ", f.value)
                                        if f.value is None:
                                            Product_Name = "----"
                                        else:
                                            Product_Name = f.value[:12]
                                    elif f.name == "timestamp":
                                        Dev_Time = f.value

                                    # Check if this frame is for the Wahoo Speed Sensor
                                    if f.name == "product_name" and f.value in speed_dev_names:
                                        Wahoo_Speed_frames += 1

                                if Dev_Charge_Str != "" and Dev_Charge_Str != "N/A":
                                    if Product_Name in GPS_Products:
                                        final_charge["GPS"] = [Dev_Time, Dev_Charge_Str]
                                    elif Product_Name in Eshift_Products:
                                        final_charge["Eshift"] = [Dev_Time, Dev_Charge_Str]
                                    else:
                                        log_file.write("{} Unknown product ({}) with battery levels {} - \n".format(
                                            str(datetime.now().strftime("%H:%M:%S")), Product_Name,
                                            Dev_Charge_Str))

            # Added 08/13/2022:
            rear_gear_changes -= 1
            front_gear_changes -= 1
#
            log_file.write("{} Frame processing completed - \n".format(str(datetime.now().strftime("%H:%M:%S"))))
            log_file.write("    {:<16} {} to {} ({})\n"
                           .format("Start/End:",
                                   start_time.replace(tzinfo=timezone.utc).astimezone(tz=None)
                                   .strftime("%b %d, %Y @ %H:%M:%S"),
                                   stop_time.replace(tzinfo=timezone.utc).astimezone(tz=None)
                                   .strftime("%H:%M:%S"),
                                   timedelta(seconds=elapsed_time_secs)))
            log_file.write("    {:<16} {:.2f} mi ({:,.2f} km) in {}\n"
                           .format("Distance:", ride_distance_meters / 1609.344, ride_distance_meters / 1000,
                                   str(timedelta(seconds=ride_time_secs))))
            log_file.write("    {:<16} Avg: {:.2f} MPH ({:.2f} KPH)  Max:  {:.2f} MPH ({:.2f} KPH)\n"
                           .format("Speed:", avg_speed_meters * 2.236936, (avg_speed_meters / 1000) * 3600,
                                   max_speed_meters * 2.236936, (max_speed_meters / 1000) * 3600))
            log_file.write("    {:<16} Max: {:.2f} ft ({:.2f} m)    Min: {:.2f} ft ({:.2f} m)\n"
                           .format("Altitude:", max_altitude_meters * 3.28084, max_altitude_meters,
                                   min_altitude_meters * 3.28084, min_altitude_meters))
            log_file.write("    {:<16} {:d} m ({:.1f}%) - Descent: {:d} m ({:.1f}%)\n"
                           .format("Ascent: ", int(total_ascent_meters), max_pos_grade_pct,
                                   int(total_descent_meters), max_neg_grade_pct))
            if rear_gear_changes > 0:
                log_file.write("    {:<16} F:{:d} / R:{:d}\n"
                               .format("Gear Changes:", front_gear_changes, rear_gear_changes))
            
            log_file.write("    {:<14}\n".format("Battery Status:"))
            # Begin building message to be sent to my phone via text
            finished_msg = "{}\n".format(stop_time.replace(tzinfo=timezone.utc).astimezone(tz=None).strftime("%b %d, %Y @ %H:%M"))
            finished_msg = finished_msg + "{:.1f} mi in {}\n".format(ride_distance_meters / 1609.344,
                                                                          str(timedelta(seconds=ride_time_secs)))
            power_msg = ""

            for key in final_charge:
                for t_key in recharge_thresholds:
                    if key == t_key[0]:
                        threshold = t_key[1]
                if int(final_charge[key][1]) < 0:
                    final_charge_msg_to = "N/A"
                else:
                    final_charge_msg_to = "{}%".format(final_charge[key][1])
                log_file.write("      {}: {}".format(Device_Tags[key], final_charge_msg_to))
                power_msg = power_msg + "{}: {}".format(Device_Tags[key], final_charge_msg_to)
                if int(final_charge[key][1]) <= threshold and int(final_charge[key][1])  >= 0:
                    log_file.write(" !RECHARGE!")
                    power_msg = power_msg + " !RECHARGE!"
                    Battery_Recharge_Flag = True
                log_file.write("\n")
                power_msg = power_msg + "\n"

            if SpeedSensor_Installed_Flag:
                log_file.write("      -- Speed: ")
                power_msg = power_msg + "Speed: "
                if Wahoo_Speed_frames == 0:
                    log_file.write("!DEAD!")
                    power_msg = power_msg + "!DEAD!"
                elif device_info_frames/10 > Wahoo_Speed_frames:
                    log_file.write("Very Low")
                    power_msg = power_msg + "Very Low"
                else:
                    log_file.write("OK")
                    power_msg = power_msg + "OK"
                log_file.write("\n")
                power_msg = power_msg + "\n"
                # print("--- power_msg ---\n",power_msg)

            if front_gear_changes < 0:
                front_gear_changes = 0
            if rear_gear_changes < 0:
                rear_gear_changes = 0

            sql = "INSERT INTO RideStats (FileName, StartTime, EndTime, RideTimeSecs, TotalTimeSecs, " \
                  "TotalDistMeters, SpeedMaxMetersSec, SpeedAvgMetersSec, AltitudeMaxMeters, AltitudeMinMeters, " \
                  "AscentMeters, AscentGradePct, DescentMeters, DescentGradePct, FrontGearChanges, " \
                  "RearGearChanges, FinalGPSCharge, FinalEshiftCharge) \
                  VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            try:
                start_time_mysql = start_time.replace(tzinfo=timezone.utc).astimezone(tz=None).strftime(
                    '%Y-%m-%d %H:%M:%S')
                stop_time_mysql = stop_time.replace(tzinfo=timezone.utc).astimezone(tz=None).strftime(
                    '%Y-%m-%d %H:%M:%S')
                insertList = [fit_file, start_time_mysql, stop_time_mysql, ride_time_secs, elapsed_time_secs,
                              ride_distance_meters, max_speed_meters, avg_speed_meters, max_altitude_meters,
                              min_altitude_meters, total_ascent_meters, max_pos_grade_pct, total_descent_meters,
                              max_neg_grade_pct, front_gear_changes, rear_gear_changes,
                              final_charge['GPS'][1], final_charge['Eshift'][1]]
                cursor.execute(sql, insertList)
                log_file.write("{} Database was updated from '{}'\n"
                               .format(str(datetime.now().strftime("%H:%M:%S")), fit_file))
            except pymysql.IntegrityError:
                log_file.write("{} Failure when adding data for: '{}'\n"
                               .format(str(datetime.now().strftime("%H:%M:%S")), fit_file))

            # Set workbook and worksheet parameters and formats

            # Open workbook with 'data_only=False' to preserve formulas previously entered
            wb = openpyxl.load_workbook(XLSX_file_name, data_only=False)
            ws = wb[worksheet_name]
            DEFAULT_FONT.name = "FreeSans"
            DEFAULT_FONT.size = 10
            YTD_Miles = 0
            YTD_Rides = 0
            # Find first empty row in target spreadsheet & build list of rides found (by date & time)
            Rides_Found = []
            # Set first empty row to 0
            sheet_row = 0
            # Starting row of scan (ie, first row of data)
            starting_row = 4
            # Last row to scan
            ending_row = 257
            # Get last used row by scanning rows
            for x in range(starting_row, ending_row):
                if ws["B{}".format(x)].value == None:
                    sheet_row = x
                    break
                else:
                    Rides_Found.append(datetime.combine(ws["B{}".format(x)].value, ws["C{}".format(x)].value))
                    # print("Row: {}, Ride Time: {}".format(x, datetime.combine(ws["B{}".format(x)].value,
                    #                                                          ws["C{}".format(x)].value)))
                    YTD_Miles += ws["E{}".format(x)].value
                    YTD_Rides += 1
            if sheet_row == 0:
                print("No empty rows found in spreadsheet")
                exit(13)

            # print("Found first empty row at #",sheet_row)
            # Check if this ride is already in the spreadsheet
            Todays_Ride_Start = datetime.combine(start_time.date(),
                                                 start_time.replace(tzinfo=timezone.utc).astimezone(tz=None).time())
            # print("This Start time: =", Todays_Ride_Start)
            if Todays_Ride_Start in Rides_Found or start_time.year != datetime.now().year:
                # print("Today's ride found in spreadsheet - no update made.")
                log_file.write("{} Today's ride found in spreadsheet or ride not in current year - no update made.\n"
                               .format(str(datetime.now().strftime("%H:%M:%S"))))
            # Write data into empty row
            else:
                ws["B{}".format(sheet_row)] = start_time.date()
                ws["B{}".format(sheet_row)].number_format = "MM/DD/YY"
                ws["C{}".format(sheet_row)] = start_time.replace(tzinfo=timezone.utc).astimezone(tz=None).time()
                ws["C{}".format(sheet_row)].number_format = "h:mm AM/PM"
                ws["D{}".format(sheet_row)] = "Synapse"
                ws["E{}".format(sheet_row)] = ride_distance_meters/1609.344
                ws["E{}".format(sheet_row)].number_format = "##0.00"
                ws["F{}".format(sheet_row)] = timedelta(seconds=ride_time_secs)
                ws["F{}".format(sheet_row)].number_format = "hh:mm:ss"
                log_file.write("{} Today's ride data written in row {} of '{}[{}]'\n"
                               .format(str(datetime.now().strftime("%H:%M:%S")), sheet_row, XLSX_file_name,
                                       worksheet_name))
                YTD_Miles += ride_distance_meters/1609.344
                YTD_Rides += 1
                # Save and Close the workbook!
                wb.save(XLSX_file_name)
            log_file.write("{} {:,} miles in {} rides year-to-date (from spreadsheet).\n"
                           .format(str(datetime.now().strftime("%H:%M:%S")), round(YTD_Miles), YTD_Rides))
        except AssertionError:
             log_file.write("Assertion error on frame #{:d} in file {}\n".format(frame_no, fit_file))

# Setup Email connection and message
port = 587  # For starttls
smtp_server = "smtp.gmail.com"
sender_email = _EMail.email_user
password = _EMail.email_password
recharge_msg = """Subject: Recharge Device(s)!

"""
context = ssl._create_unverified_context()

# Only send email/text messages if ride completed today
if datetime.now().date() == start_time.date():
    # Send text message if the list or variable has value(s)
    if len(text_msg_receiver) > 0:
        # Send 'finished' message
        message = "Bike Ride Completed!\n\n" + finished_msg
        #with smtplib.SMTP(smtp_server, port) as server:
        #    server.starttls(context=context)
        #    server.login(_EMail.email_user, _EMail.email_password)
        #    server.sendmail(_EMail.email_user, finish_email, message)
        resp = requests.post('https://textbelt.com/text', {
            'phone': text_msg_receiver,
            'message': message,
            'key': textbelt_key,
        })
        if resp.status_code == 200:
            resp_data = resp.json()
            if resp_data['success'] is True:
                log_file.write("{} 'Ride Finished' text was sent to '{}'.  Remaining quota: {} messages\n"
                               .format(str(datetime.now().strftime("%H:%M:%S")), text_msg_receiver,
                                       resp_data['quotaRemaining']))
                log_file.write("{} 'Ride Finished' message contents (between dashes):\n{}\n{}\n{}\n"
                               .format(str(datetime.now().strftime("%H:%M:%S")), dashes, message, dashes))
            else:
                log_file.write("{} 'Ride Finished' text FAILED!'.  Response was: {}\n"
                              .format(str(datetime.now().strftime("%H:%M:%S")), resp_data))
        else:
            log_file.write("{} Failure from text server!  Error code: {}"
                           .format(str(datetime.now().strftime("%H:%M:%S")), resp.status_code))

        #log_file.write("{} 'Ride Finished' text was sent to '{}'.\n"
        #               .format(str(datetime.now().strftime("%H:%M:%S")), text_msg_receiver))
        #log_file.write("{} 'Ride Finished' text server response: '{}'\n"
        #               .format(str(datetime.now().strftime("%H:%M:%S")), resp.json()))
    else:
        log_file.write("{} 'Ride Finished' text was not sent - no recipient found\n"
                       .format(str(datetime.now().strftime("%H:%M:%S"))))

    # Send email if the list or variable has value(s)
    if Battery_Recharge_Flag:
        # Send 'battery levels' email
        message = recharge_msg + power_msg + "\n\n"
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls(context=context)
            server.login(_EMail.email_user, _EMail.email_password)
            server.sendmail(_EMail.email_user, recharge_msg_to, message)
        log_file.write("{} 'Recharge' message sent to '{}'\n"
                       .format(str(datetime.now().strftime("%H:%M:%S")), recharge_msg_to))
    else:
        log_file.write("{} 'Recharge' message was not sent - No devices need to be recharged\n"
                       .format(str(datetime.now().strftime("%H:%M:%S"))))
else:
    log_file.write("{} Email & text messages NOT sent because ride did not start today\n"
                   .format(str(datetime.now().strftime("%H:%M:%S"))))
        
print("{} Completed!  {:,} Files found, {:,} Frames processed"
      .format(str(datetime.now().strftime("%H:%M:%S")), file_ctr, frame_no))
log_file.write("{} Completed!  {:,} File(s) found, {:,} Frames processed\n"
               .format(str(datetime.now().strftime("%H:%M:%S")), file_ctr, frame_no))
# Show program name and version in log
log_file.write("{} [Program {} V{}] Ended\n".format(str(datetime.now().strftime("%H:%M:%S")),
                                                    sys.argv[0], Program_Version))

# disconnect from database server
cursor.close()
db.close()
# close log file
log_file.close()
exit(0)
